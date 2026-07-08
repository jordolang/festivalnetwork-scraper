import csv
import dataclasses

from openpyxl import load_workbook

from fnscraper import export
from fnscraper.models import Event, ScoreBreakdown
from fnscraper.scoring import score_event
from tests.test_scoring import make_event


def sample():
    return [
        score_event(make_event()),
        score_event(make_event(event_id="2", name="Pipe | Fest", state="PA")),
    ]


def test_every_dataclass_field_is_exported():
    cols = export._columns()
    for f in dataclasses.fields(Event):
        assert f.name in cols
    for f in dataclasses.fields(ScoreBreakdown):
        assert f.name in cols


def test_export_all_writes_three_formats(tmp_path):
    written = export.export_all(sample(), tmp_path, "selected_shows")
    names = {p.name for p in written}
    assert names == {"selected_shows.csv", "selected_shows.md", "selected_shows.xlsx"}

    with (tmp_path / "selected_shows.csv").open() as fh:
        rows = list(csv.reader(fh))
    assert rows[0] == export._columns()
    assert len(rows) == 3
    # sorted OH before PA
    name_idx = rows[0].index("name")
    state_idx = rows[0].index("state")
    assert rows[1][state_idx] == "OH"
    assert rows[2][state_idx] == "PA"

    md = (tmp_path / "selected_shows.md").read_text()
    assert "| name |" in md.replace("| name  |", "| name |") or "name" in md.splitlines()[2]
    assert "Pipe \\| Fest" in md            # pipe escaping for table safety

    wb = load_workbook(tmp_path / "selected_shows.xlsx")
    ws = wb.active
    assert ws.max_row == 3
    assert [c.value for c in ws[1]] == export._columns()
    assert ws.freeze_panes == "A2"
    assert rows[1][name_idx]  # non-empty data


def test_infinity_and_none_become_blank(tmp_path):
    s = score_event(make_event())
    s.breakdown.cost_per_jar = float("inf")
    s.event.attendance = None
    export.export_csv([s], tmp_path / "x.csv")
    with (tmp_path / "x.csv").open() as fh:
        header, row = list(csv.reader(fh))
    assert row[header.index("cost_per_jar")] == ""
    assert row[header.index("attendance")] == ""
