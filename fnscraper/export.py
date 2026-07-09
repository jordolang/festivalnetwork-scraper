"""Full-field export of scored shows to .xlsx, .csv, and a markdown table.

Columns are generated from the dataclass definitions, so every field on
``Event`` and ``ScoreBreakdown`` is always included — new fields show up
in exports automatically.
"""

from __future__ import annotations

import csv
import dataclasses
import logging
from datetime import date
from pathlib import Path

from .models import Event, ScoreBreakdown, ScoredEvent
from .tui import sort_for_picker

log = logging.getLogger(__name__)

# Put the fields people scan for first; every remaining field follows.
LEAD_COLUMNS = [
    "name", "state", "city", "start_date", "end_date",
    "cost_per_jar", "total_cost", "booth_fee", "fuel_cost",
    "lodging_cost", "meals_cost", "est_profit", "roi", "jars_sold",
]

MONEY_FIELDS = {
    "cost_per_jar", "total_cost", "booth_fee", "fuel_cost", "lodging_cost",
    "meals_cost", "est_profit", "gross_revenue", "cogs", "exhib_fee",
    "food_fee",
}


def _columns() -> list[str]:
    all_fields = [f.name for f in dataclasses.fields(Event)]
    all_fields += [f.name for f in dataclasses.fields(ScoreBreakdown)]
    ordered = [c for c in LEAD_COLUMNS if c in all_fields]
    ordered += [c for c in all_fields if c not in ordered]
    return ordered


def _cell(value):
    if value is None:
        return ""
    if isinstance(value, float):
        if value == float("inf"):
            return ""
        return round(value, 2)
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, date):
        return value.isoformat()
    return value


def _rows(scored: list[ScoredEvent]) -> tuple[list[str], list[list]]:
    cols = _columns()
    rows = []
    for s in sort_for_picker(scored):
        merged = {**dataclasses.asdict(s.event), **dataclasses.asdict(s.breakdown)}
        rows.append([_cell(merged.get(c)) for c in cols])
    return cols, rows


def export_csv(scored: list[ScoredEvent], path: Path) -> Path:
    cols, rows = _rows(scored)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(cols)
        w.writerows(rows)
    return path


def export_markdown(scored: list[ScoredEvent], path: Path) -> Path:
    cols, rows = _rows(scored)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Selected shows — full data",
        "",
        "| " + " | ".join(cols) + " |",
        "|" + "|".join("---" for _ in cols) + "|",
    ]
    for row in rows:
        cells = [str(c).replace("|", "\\|").replace("\n", " ") for c in row]
        lines.append("| " + " | ".join(cells) + " |")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def export_xlsx(scored: list[ScoredEvent], path: Path) -> Path | None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed; skipping %s "
                    "(pip install openpyxl)", path)
        return None

    cols, rows = _rows(scored)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Selected shows"
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for idx, col in enumerate(cols, 1):
        letter = get_column_letter(idx)
        if col in MONEY_FIELDS:
            for cell in ws[letter][1:]:
                cell.number_format = '"$"#,##0.00'
        width = max(len(col), *(len(str(r[idx - 1])) for r in rows)) if rows else len(col)
        ws.column_dimensions[letter].width = min(max(width + 2, 8), 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return path


def export_all(scored: list[ScoredEvent], out_dir: Path, stem: str) -> list[Path]:
    """Write ``<stem>.csv``, ``<stem>.md``, and ``<stem>.xlsx``."""
    out_dir.mkdir(parents=True, exist_ok=True)
    written = [
        export_csv(scored, out_dir / f"{stem}.csv"),
        export_markdown(scored, out_dir / f"{stem}.md"),
    ]
    xlsx = export_xlsx(scored, out_dir / f"{stem}.xlsx")
    if xlsx:
        written.append(xlsx)
    return written
