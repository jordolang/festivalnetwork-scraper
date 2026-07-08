"""Import previously exported show lists (.csv, .xlsx, or .md table).

The export format is column-per-dataclass-field, so imports reconstruct
full ``ScoredEvent`` objects.  Type coercion is derived from the dataclass
annotations — a field annotated ``int | None`` gets int parsing, ``date``
gets ISO-date parsing, and so on.
"""

from __future__ import annotations

import csv
import dataclasses
import logging
from datetime import date
from pathlib import Path

from .models import Event, ScoreBreakdown, ScoredEvent

log = logging.getLogger(__name__)

_EVENT_FIELDS = {f.name: str(f.type) for f in dataclasses.fields(Event)}
_BREAKDOWN_FIELDS = {f.name: str(f.type) for f in dataclasses.fields(ScoreBreakdown)}


def _coerce(value, annotation: str):
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return [] if "list" in annotation else None
    if "list" in annotation:
        return [p.strip() for p in text.split(";") if p.strip()]
    if "bool" in annotation:
        if isinstance(value, bool):
            return value
        return text.lower() in ("true", "1", "yes", "x")
    if "date" in annotation:
        if isinstance(value, date):
            return value
        return date.fromisoformat(text[:10])
    if "int" in annotation:
        return int(float(text.replace(",", "")))
    if "float" in annotation:
        return float(text.replace(",", "").replace("$", ""))
    return text


def _row_to_scored(record: dict) -> ScoredEvent:
    ev_kwargs, bd_kwargs = {}, {}
    for key, raw in record.items():
        if key in _EVENT_FIELDS:
            try:
                ev_kwargs[key] = _coerce(raw, _EVENT_FIELDS[key])
            except (ValueError, TypeError):
                log.debug("could not coerce event field %s=%r", key, raw)
        elif key in _BREAKDOWN_FIELDS:
            try:
                val = _coerce(raw, _BREAKDOWN_FIELDS[key])
                if val is not None:
                    bd_kwargs[key] = val
            except (ValueError, TypeError):
                log.debug("could not coerce breakdown field %s=%r", key, raw)
    if not ev_kwargs.get("event_id") or not ev_kwargs.get("url"):
        raise ValueError(
            f"row is missing event_id/url (got: {list(record)[:6]}...); "
            "import files must come from this tool's exports"
        )
    ev_kwargs.setdefault("name", "")
    return ScoredEvent(event=Event(**ev_kwargs), breakdown=ScoreBreakdown(**bd_kwargs))


def _read_csv(path: Path) -> list[dict]:
    with path.open(newline="") as fh:
        return list(csv.DictReader(fh))


def _read_xlsx(path: Path) -> list[dict]:
    from openpyxl import load_workbook
    ws = load_workbook(path, read_only=True).active
    rows = ws.iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in next(rows)]
    return [dict(zip(header, row)) for row in rows if any(c is not None for c in row)]


def _split_md_row(line: str) -> list[str]:
    import re
    cells = re.split(r"(?<!\\)\|", line.strip().strip("|"))
    return [c.strip().replace("\\|", "|") for c in cells]


def _read_markdown(path: Path) -> list[dict]:
    records = []
    header: list[str] | None = None
    for line in path.read_text().splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = _split_md_row(line)
        if header is None:
            header = cells
            continue
        if set("".join(cells)) <= {"-", " ", ":"}:   # separator row
            continue
        records.append(dict(zip(header, cells)))
    return records


def load_shows(path: str | Path) -> list[ScoredEvent]:
    """Load an exported show list; format chosen by file extension."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        records = _read_csv(path)
    elif suffix == ".xlsx":
        records = _read_xlsx(path)
    elif suffix in (".md", ".markdown"):
        records = _read_markdown(path)
    else:
        raise ValueError(f"unsupported import format: {suffix} "
                         "(use .csv, .xlsx, or .md)")
    shows = [_row_to_scored(r) for r in records]
    log.info("imported %d show(s) from %s", len(shows), path)
    return shows
